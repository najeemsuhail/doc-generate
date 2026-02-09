import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Customers"

# Add headers
headers = ['CUSTOMER NAME', 'Address', 'City', 'State', 'Zip', 'Status', 'Outstanding_Amount']
ws.append(headers)

# Style header row
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add sample customer data
sample_data = [
    ['John Smith', '123 Main Street', 'Springfield', 'IL', '62701', 'Overdue', 1500.00],
    ['Sarah Johnson', '456 Oak Avenue', 'Chicago', 'IL', '60601', 'Delinquent', 2500.00],
    ['Michael Brown', '789 Pine Road', 'Peoria', 'IL', '61602', 'Pending', 750.50],
    ['Emily Davis', '321 Elm Street', 'Naperville', 'IL', '60540', 'Active', 0.00],
    ['Robert Wilson', '654 Maple Drive', 'Aurora', 'IL', '60506', 'Overdue', 3200.75],
    ['Jennifer Garcia', '987 Cedar Lane', 'Rockford', 'IL', '61101', 'Active', 0.00],
    ['David Martinez', '147 Birch Boulevard', 'Joliet', 'IL', '60432', 'Delinquent', 1800.00],
    ['Lisa Anderson', '258 Spruce Street', 'Champaign', 'IL', '61820', 'Pending', 500.00],
]

for row in sample_data:
    ws.append(row)

# Adjust column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 18

# Save the file
wb.save('customers.xlsx')
print("✓ Sample Excel file 'customers.xlsx' created with 8 sample customers!")
